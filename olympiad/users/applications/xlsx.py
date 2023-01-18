import datetime

from django.http import HttpResponse
from django.utils.encoding import escape_uri_path
from openpyxl import load_workbook, Workbook

from users.applications.create_identiier import CreateIdentifier
from users.models import Apps


def UploadAppsFromExcel(file, id_user) -> dict:
    """Добавление заявок из загруженного xlsx файла"""
    error = ''
    wb = load_workbook(file)
    sheet = wb.active
    cols_dict = {
        'surname': 0,
        'name': 0,
        'age': 0,
        'study_year': 0,
        'study_kind': 0,
        'study_duration': 0,
        'teacher': 0,
        'group': 0
    }
    for i, col in enumerate(sheet.columns, start=1):
        if sheet.cell(row=1, column=i).value == 'Фамилия':
            cols_dict['surname'] = i
        if sheet.cell(row=1, column=i).value == 'Имя':
            cols_dict['name'] = i
        if sheet.cell(row=1, column=i).value == 'Возраст':
            cols_dict['age'] = i
        if sheet.cell(row=1, column=i).value == 'Год обучения':
            cols_dict['study_year'] = i
        if sheet.cell(row=1, column=i).value == 'Вид образования':
            cols_dict['study_kind'] = i
        if sheet.cell(row=1, column=i).value == 'Срок реализации программы обучения':
            cols_dict['study_duration'] = i
        if sheet.cell(row=1, column=i).value == 'ФИО преподавателя':
            cols_dict['teacher'] = i
        if sheet.cell(row=1, column=i).value == 'Группа':
            cols_dict['group'] = i
    for key, value in cols_dict.items():
        if value == 0:
            if key == 'surname':
                error += 'Не найден столбец "Фамилия"; '
            elif key == 'name':
                error += 'Не найден столбец "Имя"; '
            elif key == 'age':
                error += 'Не найден столбец "Возраст"; '
            elif key == 'study_year':
                error += 'Не найден столбец "Год обучения"; '
            elif key == 'study_kind':
                error += 'Не найден столбец "Вид образования"; '
            elif key == 'study_duration':
                error += 'Не найден столбец "Срок реализации программы обучения"; '
            elif key == 'teacher':
                error += 'Не найден столбец "ФИО преподавателя"; '
            else:
                error += 'Не найден столбец "Группа"; '
    if len(error) > 0:
        return {
            'error': error
        }
    apps_count = Apps.objects.filter(contact_person_id=id_user).count()
    for i, row in enumerate(sheet.rows, start=1):
        if i != 1:
            correct = True
            if sheet.cell(row=i, column=cols_dict['study_kind']).value not in [
                'Общее',
                'Дополнительное',
                'Профессиональное'
            ]:
                correct = False
            try:
                new_app = Apps(
                    date_create=datetime.date.today(),
                    contact_person_id=id_user,
                    surname=sheet.cell(row=i, column=cols_dict['surname']).value,
                    name=sheet.cell(row=i, column=cols_dict['name']).value,
                    age=int(sheet.cell(row=i, column=cols_dict['age']).value),
                    study_year=int(sheet.cell(row=i, column=cols_dict['study_year']).value),
                    study_kind=sheet.cell(row=i, column=cols_dict['study_kind']).value,
                    study_duration=int(sheet.cell(row=i, column=cols_dict['study_duration']).value),
                    teacher=sheet.cell(row=i, column=cols_dict['teacher']).value,
                    group=sheet.cell(row=i, column=cols_dict['group']).value,
                    identifier=CreateIdentifier()
                )
            except BaseException:
                correct = False
            if correct is False:
                error += f'Ошибка при создании заявки в строке №{str(i+1)}'
            else:
                new_app.save()
    if len(error) > 0:
        if apps_count != Apps.objects.filter(contact_person_id=id_user).count():
            return {
                'error': f'Заявки загружены не полностью: {error}'
            }
        else:
            return {
                'error': error
            }
    return {
        'success': 'Заявки успешно загружены'
    }


def DownloadApps(id_user, search_str):
    """Формирование Response с xlsx файлом заявок с поисковыми критериями пользователя"""
    filename = 'Заявки.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="' + escape_uri_path(filename) + '"'
    apps = Apps.objects.filter(contact_person_id=id_user).order_by('-date_create')
    if len(search_str) > 0:
        search_arr = search_str.partition('&')
        for search in search_arr:
            if 'data_create' in search:
                apps = apps.filter(date_create=search[search.find('=')+1:])
            if 'surname' in search:
                apps = apps.filter(surname__contains=search[search.find('=')+1:])
            if 'name' in search:
                apps = apps.filter(name__contains=search[search.find('=')+1:])
            if 'age' in search:
                apps = apps.filter(age=int(search[search.find('=')+1:]))
            if 'study_year' in search:
                apps = apps.filter(study_year=int(search[search.find('=')+1:]))
            if 'study_kind' in search:
                apps = apps.filter(study_kind=search[search.find('=')+1:])
            if 'teacher' in search:
                apps = apps.filter(teacher__contains=search[search.find('=')+1:])
            if 'study_duration' in search:
                apps = apps.filter(study_duration=int(search[search.find('=')+1:]))
            if 'group' in search:
                apps = apps.filter(group__contains=search[search.find('=')+1:])
            if 'identifier' in search:
                apps = apps.filter(identifier__contains=search[search.find('=')+1:])
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Заявки'
    cell = worksheet.cell(row=1, column=1)
    cell.value = 'Дата подачи заявки'
    cell = worksheet.cell(row=1, column=2)
    cell.value = 'Фамилия'
    cell = worksheet.cell(row=1, column=3)
    cell.value = 'Имя'
    cell = worksheet.cell(row=1, column=4)
    cell.value = 'Возраст'
    cell = worksheet.cell(row=1, column=5)
    cell.value = 'Год обучения'
    cell = worksheet.cell(row=1, column=6)
    cell.value = 'Вид образования'
    cell = worksheet.cell(row=1, column=7)
    cell.value = 'Срок реализации программы обучения'
    cell = worksheet.cell(row=1, column=8)
    cell.value = 'ФИО преподавателя'
    cell = worksheet.cell(row=1, column=9)
    cell.value = 'Группа'
    cell = worksheet.cell(row=1, column=10)
    cell.value = 'Идентификатор'
    for i, app in enumerate(apps, start=1):
        cell = worksheet.cell(row=i+1, column=1)
        cell.value = app.date_create.strftime('%d.%m.%Y')
        cell = worksheet.cell(row=i+1, column=2)
        cell.value = app.surname
        cell = worksheet.cell(row=i+1, column=3)
        cell.value = app.name
        cell = worksheet.cell(row=i+1, column=4)
        cell.value = str(app.age)
        cell = worksheet.cell(row=i+1, column=5)
        cell.value = str(app.study_year)
        cell = worksheet.cell(row=i+1, column=6)
        cell.value = app.study_kind
        cell = worksheet.cell(row=i+1, column=7)
        cell.value = str(app.study_duration)
        cell = worksheet.cell(row=i+1, column=8)
        cell.value = app.teacher
        cell = worksheet.cell(row=i+1, column=9)
        cell.value = app.group
        cell = worksheet.cell(row=i+1, column=10)
        cell.value = app.identifier
    workbook.save(response)
    return response
