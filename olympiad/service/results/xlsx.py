from django.http import HttpResponse
from django.utils.encoding import escape_uri_path
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from service.models import Olympiads, OlympiadsLevels, ResultsSessions, ResultsQuestions, Questions
from users.models import Apps


def DownloadResult(olympiad_id):
    """Формирование Response с xlsx файлом результатов олимпиады"""
    filename = 'Результаты олимпиады.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="' + escape_uri_path(filename) + '"'
    olympiad = Olympiads.objects.get(id=olympiad_id)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Результат'
    cell = worksheet.cell(row=1, column=1)
    cell.value = 'Тема олимпиады:'
    cell = worksheet.cell(row=1, column=2)
    cell.value = olympiad.theme
    cell = worksheet.cell(row=2, column=1)
    cell.value = 'Дата проведения:'
    cell = worksheet.cell(row=2, column=2)
    cell.value = olympiad.event_date.strftime('%d.%m.%Y г.')
    cell = worksheet.cell(row=3, column=1)
    cell.value = 'Зарегистрировано участников:'
    cell = worksheet.cell(row=3, column=2)
    cell.value = str(Apps.objects.filter(date_create__gte=olympiad.date_reg_start).\
                     filter(date_create__lte=olympiad.date_reg_end).count())
    cell = worksheet.cell(row=4, column=1)
    cell.value = 'Время на прохождение :'
    cell = worksheet.cell(row=4, column=2)
    cell.value = f'{olympiad.time_complete} минут'
    worksheet.merge_cells('A6:A7')
    cell = worksheet.cell(row=6, column=1)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Образовательная организация'
    worksheet.merge_cells('B6:B7')
    cell = worksheet.cell(row=6, column=2)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Участник'
    levels = OlympiadsLevels.objects.filter(olympiad_id=olympiad.id).order_by('seq_number')
    column_letter = get_column_letter(2+levels.count())
    merged = f'C6:{column_letter}6'
    worksheet.merge_cells(merged)
    cell = worksheet.cell(row=6, column=3)
    cell.alignment = Alignment(horizontal='center')
    cell.value = "Баллы за ответы участника (по темам)"
    worksheet.merge_cells(f'{get_column_letter(3+levels.count())}6:{get_column_letter(3+levels.count())}7')
    cell = worksheet.cell(row=6, column=3+levels.count())
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Итоговый балл'
    for index, level in enumerate(levels, start=1):
        cell = worksheet.cell(row=7, column=2+index)
        cell.alignment = Alignment(horizontal='center')
        cell.value = level.level.level
    sessions = ResultsSessions.objects.filter(olympiad_theme=olympiad.theme)
    for row, session in enumerate(sessions, start=1):
        cell = worksheet.cell(row=7+row, column=1)
        cell.alignment = Alignment(horizontal='center')
        cell.value = session.participant_oo
        cell = worksheet.cell(row=7+row, column=2)
        cell.alignment = Alignment(horizontal='center')
        cell.value = f'{session.participant_surname} {session.participant_name}'
        answers = ResultsQuestions.objects.filter(result_session_id=session.id)
        total_count = 0
        for col, level in enumerate(levels, start=1):
            count = 0
            for answer in answers:
                if Questions.objects.get(id=answer.question_id).level_id == level.level_id:
                    count += answer.points
            total_count += count
            cell = worksheet.cell(row=7+row, column=2+col)
            cell.alignment = Alignment(horizontal='center')
            cell.value = str(count)
        cell = worksheet.cell(row=7+row, column=3+levels.count())
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.value = str(total_count)
    workbook.save(response)
    return response
